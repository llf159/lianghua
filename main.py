"""Unified back‑testing entry script.

Usage
-----
# 批量回测整个 DATA_DIR
python main.py

# 仅回测指定股票
python main.py --code 000001
"""

import argparse
import glob
import os
import sys
from datetime import datetime
from concurrent.futures import ProcessPoolExecutor, as_completed
from utils import ensure_datetime_index
import pandas as pd
from tqdm import tqdm
from tabulate import tabulate
from backtest_core import backtest
from strategy_buy import buy_signal
from strategy_sell import sell_signal
from config import (
    DATA_DIR,
    HOLD_DAYS,
    BUY_MODE,
    SELL_MODE,
    MAX_HOLD_DAYS,
    FALLBACK_SELL_MODE,
    START_DATE,
    END_DATE,
)

N_WORKERS = max(1, os.cpu_count() - 1)

def load_file_by_code(code: str) -> str:
    """根据股票代码查找文件路径。"""
    pattern = os.path.join(DATA_DIR, f"*{code}*.csv")
    matches = glob.glob(pattern)
    if not matches:
        raise FileNotFoundError(f"找不到股票 {code} 的历史数据文件：{pattern}")
    # 如果有多于一个匹配，取文件名最短/最早的一个
    return sorted(matches)[0]


def run_backtest_on_file(file_path: str) -> tuple[str, dict]:
    """子进程回测函数。"""
    df = pd.read_csv(file_path)
    df = ensure_datetime_index(df)
    df = df[(df.index >= START_DATE) & (df.index <= END_DATE)]

    summary, _ = backtest(
        df,
        HOLD_DAYS=HOLD_DAYS,
        BUY_MODE=BUY_MODE,
        SELL_MODE=SELL_MODE,
        MAX_HOLD_DAYS=MAX_HOLD_DAYS,
        FALLBACK_SELL_MODE=FALLBACK_SELL_MODE,
        buy_signal=buy_signal,
        sell_signal=sell_signal,
        record_trades=False,
    )
    code = os.path.splitext(os.path.basename(file_path))[0]
    return code, summary


def run_single(code: str) -> None:
    file_path = load_file_by_code(code)
    df = pd.read_csv(file_path, parse_dates=['trade_date'])
    df = ensure_datetime_index(df)
    df = df[(df.index >= START_DATE) & (df.index <= END_DATE)]

    summary, trades = backtest(
        df,
        HOLD_DAYS=HOLD_DAYS,
        BUY_MODE=BUY_MODE,
        SELL_MODE=SELL_MODE,
        MAX_HOLD_DAYS=MAX_HOLD_DAYS,
        FALLBACK_SELL_MODE=FALLBACK_SELL_MODE,
        buy_signal=buy_signal,
        sell_signal=sell_signal,
        record_trades=True,
    )

    print("\n===== 概览 =====")
    for k, v in summary.items():
        print(f"{k}: {v}")

    if trades:
        trades_df = pd.DataFrame(trades)
        trades_df.columns = ["Buy Date", "Sell Date", "Hold Days", "Buy Price", "Sell Price", "Return Rate"]
        print("\n===== 交易明细 =====")
        print(tabulate(trades_df, headers='keys', tablefmt='github', showindex=False))
    else:
        print("\n没有产生任何交易信号。")


def run_batch() -> None:
    files = glob.glob(os.path.join(DATA_DIR, "*.csv"))
    if not files:
        print(f"DATA_DIR = {DATA_DIR} 下没有找到任何数据文件！", file=sys.stderr)
        sys.exit(1)

    results: list[tuple[str, dict]] = []

    with ProcessPoolExecutor(max_workers=N_WORKERS) as executor:
        futures = {executor.submit(run_backtest_on_file, f): f for f in files}
        for fut in tqdm(as_completed(futures), total=len(futures), desc="回测进度"):
            code, summary = fut.result()
            summary["代码"] = code
            results.append((code, summary))

    # 汇总结果
    results = [dict(filename=c + ".csv", **{k: v for k, v in s.items() if k != "代码"}) for c, s in results]
    df = pd.DataFrame(results)

    # === 类型转换处理 ===
    valid_rows = df[pd.to_numeric(df["盈亏比"], errors='coerce').notna()]
    valid_rows.loc[:, "信号数"] = pd.to_numeric(valid_rows["信号数"], errors='coerce')
    valid_rows.loc[:, "盈亏比"] = pd.to_numeric(valid_rows["盈亏比"], errors='coerce')

    total_signals = valid_rows["信号数"].sum()
    if total_signals > 0:
        weighted_pl_ratio = (valid_rows["盈亏比"] * valid_rows["信号数"]).sum() / total_signals
    else:
        weighted_pl_ratio = 0

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = "results"
    os.makedirs(out_dir, exist_ok=True)
    xlsx_path = os.path.join(out_dir, f"summary_{timestamp}.xlsx")
    df.to_excel(xlsx_path, index=False)

    # === 插入加权平均盈亏比 ===
    import openpyxl
    from openpyxl.styles import PatternFill
    from openpyxl.formatting.rule import CellIsRule

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    # G列胜率、H列平均盈亏比标红
    max_row = ws.max_row
    ws.conditional_formatting.add(f"G2:G{max_row}", CellIsRule(operator="lessThan", formula=["1"], fill=red_fill))
    ws.conditional_formatting.add(f"H2:H{max_row}", CellIsRule(operator="lessThan", formula=["1"], fill=red_fill))

    ws.insert_rows(1)
    ws.cell(row=1, column=1, value="加权平均盈亏比")
    ws.cell(row=1, column=2, value=round(weighted_pl_ratio, 4))

    wb.save(xlsx_path)

    print(f"\n已保存汇总结果（含加权平均盈亏比）→ {xlsx_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="量化回测入口脚本")
    parser.add_argument("--code", help="只回测指定股票代码，例如 000001")

    args = parser.parse_args()

    if args.code:
        run_single(args.code)
    else:
        run_batch()


if __name__ == "__main__":
    main()
