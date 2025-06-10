import pandas as pd
import numpy as np
import argparse
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

from strategy import buy_signal, should_sell_open, should_sell_close
from config import HOLD_DAYS, START_DATE, END_DATE, BUY_MODE, SELL_MODE, DATA_DIR, MAX_HOLD_DAYS
from utils import ensure_datetime_index

# === 初始化运行时间 & 文件路径 ===
RUN_TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M")
RESULTS_DIR = "results"
os.makedirs(RESULTS_DIR, exist_ok=True)
CSV_PATH = os.path.join(RESULTS_DIR, f"summary_{RUN_TIMESTAMP}.csv")
XLSX_PATH = CSV_PATH.replace(".csv", ".xlsx")


def backtest(df, n_days, buy_mode):
    signals = buy_signal(df)
    buy_dates = df.index[signals]

    results, wins, losses = [], [], []

    for buy_date in buy_dates:
        idx = df.index.get_loc(buy_date)

        if buy_mode == 'open':
            if idx + 1 + n_days < len(df):
                prev_close = df.iloc[idx]['close']
                open_price = df.iloc[idx + 1]['open']
                limit_up_price = round(prev_close * 1.10, 2)
                if open_price >= limit_up_price:
                    continue
                buy_price = open_price
                sell_idx = idx + 1 + n_days
            else:
                continue

        elif buy_mode == 'close':
            if idx + n_days < len(df):
                buy_price = df.iloc[idx]['close']
                sell_idx = idx + n_days
            else:
                continue
        else:
            raise ValueError(f"不支持的买入模式: {buy_mode}")

        if SELL_MODE == 'open':
            sell_price = df.iloc[sell_idx]['open']
        else:
            sell_price = df.iloc[sell_idx]['close']

        ret = (sell_price - buy_price) / buy_price
        results.append(ret)
        if ret > 0:
            wins.append(ret)
        elif ret < 0:
            losses.append(ret)

    returns = np.array(results)
    win_avg = np.mean(wins) if wins else 0
    loss_avg = abs(np.mean(losses)) if losses else 0
    win_rate = (returns > 0).mean() if len(results) else 0
    avg_return = returns.mean() if len(results) else 0

    if not losses:
        pl_ratio = avg_return
    else:
        pl_ratio = win_avg / loss_avg if loss_avg else 0

    score = pl_ratio * win_rate

    summary = {
        "信号数": len(results),
        "平均收益": f"{avg_return:.2%}" if len(results) else "N/A",
        "胜率": f"{win_rate:.2%}" if len(results) else "N/A",
        "最大收益": f"{returns.max():.2%}" if len(results) else "N/A",
        "最大亏损": f"{returns.min():.2%}" if len(results) else "N/A",
        "盈亏比": round(pl_ratio, 3) if len(results) else None,
        "平均盈亏比": round(score, 3) if len(results) else None
    }

    return summary


def run_backtest_on_file(file_path):
    df = pd.read_csv(file_path, parse_dates=['trade_date'])
    df.rename(columns={'trade_date': 'date'}, inplace=True)
    df = ensure_datetime_index(df)
    df = df[(df.index >= START_DATE) & (df.index <= END_DATE)]
    summary = backtest(df, HOLD_DAYS, BUY_MODE)
    return os.path.basename(file_path), summary


def main():
    parser = argparse.ArgumentParser(description="量化策略回测工具 v3")
    parser.add_argument('--file', help='单个CSV数据文件路径（可选）')
    args = parser.parse_args()

    all_results = []

    if args.file:
        files = [args.file]
    else:
        if not os.path.exists(DATA_DIR):
            print(f"配置中的数据目录不存在: {DATA_DIR}")
            return
        files = [os.path.join(DATA_DIR, f) for f in os.listdir(DATA_DIR) if f.endswith('.csv')]

    for fpath in files:
        fname, summary = run_backtest_on_file(fpath)
        print(f"【{fname}】结果：")
        for k, v in summary.items():
            print(f"  {k}: {v}")
        all_results.append(dict(filename=fname, **summary))

    # 保存 CSV/Excel 结果
    df_all = pd.DataFrame(all_results)
    df_all.to_csv(CSV_PATH, index=False, encoding='utf-8-sig')
    df_all.to_excel(XLSX_PATH, index=False)

    # 添加条件格式
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    ws.conditional_formatting.add("G2:G6000", FormulaRule(formula=["AND(ISNUMBER(G2), G2<1)"], fill=red_fill))
    ws.conditional_formatting.add("H2:H6000", FormulaRule(formula=["AND(ISNUMBER(H2), H2<1)"], fill=red_fill))
    ws.conditional_formatting.add("H2", FormulaRule(formula=["COUNTIF(H3:H1000,\"<1\")>0"], fill=red_fill))
    wb.save(XLSX_PATH)


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print("程序发生错误：", e)
        input("按下回车关闭窗口...")
