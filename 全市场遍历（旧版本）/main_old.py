import pandas as pd
import numpy as np
import argparse
import os
from strategy import buy_signal
from config import HOLD_DAYS, START_DATE, END_DATE, BUY_MODE
from utils import ensure_datetime_index
from datetime import datetime
# 程序启动时的时间戳，只生成一次
RUN_TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M")



def backtest(df, n_days,buy_mode): 
    signals = buy_signal(df)
    buy_dates = df.index[signals]

    results = []
    wins = []
    losses = []
    
    
    for buy_date in buy_dates:
        idx = df.index.get_loc(buy_date)

        if buy_mode == 'open':
            if idx + 1 >= len(df):
                continue
            buy_price = df.iloc[idx + 1]['open']
            buy_idx = idx + 1
            sell_start_idx = idx + 2  # 第二天开始才有持仓

        elif buy_mode == 'close':
            if idx >= len(df):
                continue
            buy_price = df.iloc[idx]['close']
            buy_idx = idx
            sell_start_idx = idx + 1  # 第二天才能卖出

        elif buy_mode == 'signal_open':
            if idx >= len(df):
                continue
            buy_price = df.iloc[idx]['open']
            buy_idx = idx
            sell_start_idx = idx + 1

        else:
            raise ValueError(f"不支持的买入模式: {buy_mode}")

        ret = (sell_price - buy_price) / buy_price
        results.append(ret)

        if ret > 0:
            wins.append(ret)
        elif ret < 0:
            losses.append(ret)

    #统计部分
    returns = np.array(results)
    win_avg = np.mean(wins) if wins else 0
    loss_avg = abs(np.mean(losses)) if losses else 0
    win_rate = (returns > 0).mean() if len(results) else 0
    avg_return = returns.mean() if len(results) else 0

    # 盈亏比逻辑
    if not losses:
        pl_ratio = np.nan
        corrected_pl = avg_return
        score = corrected_pl * win_rate
    else:
        pl_ratio_val = win_avg / loss_avg if loss_avg else 0
        pl_ratio = f"{pl_ratio_val:.2f}"
        corrected_pl = pl_ratio_val
        score = corrected_pl * win_rate
    
    summary = {
        "信号数": len(results),
        "平均收益": f"{returns.mean():.2%}" if len(returns) else "N/A",
        "胜率": f"{(returns > 0).mean():.2%}" if len(returns) else "N/A",
        "最大收益": f"{returns.max():.2%}" if len(returns) else "N/A",
        "最大亏损": f"{returns.min():.2%}" if len(returns) else "N/A",
        "盈亏比": float(f"{corrected_pl:.3f}") if pl_ratio == "all win" else float(pl_ratio),
        "平均盈亏比": float(f"{score:.3f}") if len(results) else None

    }

    return summary

#文件写入部分
def run_backtest_on_file(file_path):
    df = pd.read_csv(file_path, parse_dates=['trade_date'])
    df.rename(columns={'trade_date': 'date'}, inplace=True)
    df = ensure_datetime_index(df)

    df = df[(df.index >= START_DATE) & (df.index <= END_DATE)]

    summary = backtest(df, HOLD_DAYS,buy_mode=BUY_MODE)
    print(f"【{os.path.basename(file_path)}】结果：")
    for k, v in summary.items():
        print(f"  {k}: {v}")

    return os.path.basename(file_path), summary

from config import DATA_DIR  # 放到顶部导入区

def main():
    parser = argparse.ArgumentParser(description="量化策略回测工具 v2")
    parser.add_argument('--file', help='单个CSV数据文件路径（可选）')
    args = parser.parse_args()

    results = []

    if args.file:
        print("文件名                 | 信号数 | 平均收益 | 胜率   | 最大收益 | 最大亏损")
        print("-" * 70)
        # 回测指定单个文件
        fname, summary = run_backtest_on_file(args.file)
        results.append((fname, summary))
    else:
        # 回测配置目录中的所有文件
        if not os.path.exists(DATA_DIR):
            print(f"配置中的数据目录不存在: {DATA_DIR}")
            return

        for fname in os.listdir(DATA_DIR):
            if fname.endswith('.csv'):
                fpath = os.path.join(DATA_DIR, fname)
                fname, summary = run_backtest_on_file(fpath)
                results.append((fname, summary))
                print(f"{fname:<20} | {summary['信号数']:>3} | {summary['平均收益']:>6} | {summary['胜率']:>6} | {summary['最大收益']:>6} | {summary['最大亏损']:>6}")

                today_str = datetime.today().strftime('%Y-%m-%d_%H-%M')
                output_path = f"results/summary_{today_str}.csv"
                write_header = not os.path.exists(output_path)
                df_partial = pd.DataFrame([dict(filename=fname, **summary)])
                import openpyxl
                from openpyxl.styles import PatternFill
                from openpyxl.formatting.rule import CellIsRule, FormulaRule

                # 写入 CSV（保留原逻辑）
                df_partial.to_csv(output_path, mode='a', header=write_header, index=False, encoding="utf-8-sig")

                # ==== 写入 Excel 并加条件格式 ====
                xlsx_path = output_path.replace(".csv", ".xlsx")

                if write_header:
                    df_all = pd.DataFrame([dict(filename=fname, **summary)])
                    df_all.to_excel(xlsx_path, index=False)
                else:
                    # 如果已存在，读取再合并
                    df_old = pd.read_excel(xlsx_path)
                    df_all = pd.concat([df_old, pd.DataFrame([dict(filename=fname, **summary)])], ignore_index=True)
                    df_all.to_excel(xlsx_path, index=False)

                # 打开 Excel 添加条件格式
                wb = openpyxl.load_workbook(xlsx_path)
                ws = wb.active

                red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

                # 假设列顺序：A 文件名 | B 信号数 | C 平均收益 | D 胜率 | E 最大收益 | F 最大亏损 | G 盈亏比 | H 策略评分

                # 1. 盈亏比 < 1 标红
                ws.conditional_formatting.add("G2:G1000", FormulaRule(formula=["AND(ISNUMBER(G2), G2<1)"], fill=red_fill))
                # 2. 平均盈亏比 < 1 标红
                ws.conditional_formatting.add("H2:H1000", FormulaRule(formula=["AND(ISNUMBER(H2), H2<1)"], fill=red_fill))
                # 3. H2 表头：如果 H 列有任何 <1，则标红提醒
                ws.conditional_formatting.add("H2", FormulaRule(formula=["COUNTIF(H3:H1000,\"<1\")>0"], fill=red_fill))
                
                wb.save(xlsx_path)


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print("程序发生错误：", e)
        input("按下回车关闭窗口...")