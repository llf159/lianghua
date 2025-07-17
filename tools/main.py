import pandas as pd
import numpy as np
import argparse
import time
import os
import openpyxl
import sys

from concurrent.futures import ProcessPoolExecutor, as_completed, ThreadPoolExecutor
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from tqdm import tqdm
from strategy import buy_signal, sell_signal
from config import HOLD_DAYS, START_DATE, END_DATE, BUY_MODE, SELL_MODE, DATA_DIR, FALLBACK_SELL_MODE, MAX_HOLD_DAYS
from utils import ensure_datetime_index

#初始化运行时间 & 文件路径
sys.stdout.reconfigure(encoding='utf-8')
RUN_TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
RESULTS_DIR = "results"
os.makedirs(RESULTS_DIR, exist_ok=True)
CSV_PATH = os.path.join(RESULTS_DIR, f"summary_{RUN_TIMESTAMP}.csv")
XLSX_PATH = CSV_PATH.replace(".csv", ".xlsx")

#逻辑部分#################################################################################################
def backtest(df, *, HOLD_DAYS, BUY_MODE, SELL_MODE, MAX_HOLD_DAYS, FALLBACK_SELL_MODE, buy_signal, sell_signal):
    signals = buy_signal(df)
    buy_dates = pd.to_datetime(df.index[signals])
    results, wins, losses = [], [], []

    holding_until = None

    for buy_date in buy_dates:
        if holding_until and buy_date <= holding_until:
            continue
        
        idx = df.index.get_loc(buy_date)

        # === 买入逻辑 ===
        if BUY_MODE == 'open':
            if idx + 1 < len(df):
                prev_close = df.iloc[idx]['close']
                open_price = df.iloc[idx + 1]['open']
                limit_up_price = round(prev_close * 1.099, 2)
                if open_price >= limit_up_price:                          # 如果开盘价大于等于涨停价，则不买入
                    continue
                buy_price = open_price
                sell_start_idx = idx + 1
            else:
                continue

        elif BUY_MODE == 'close':
            if idx + 1 < len(df):
                buy_price = df.iloc[idx]['close']
                sell_start_idx = idx + 1
            else:
                continue

        elif BUY_MODE == 'signal_open': 
            if SELL_MODE in ('open', 'close') and idx + HOLD_DAYS >= len(df):
                continue
            buy_price = df.iloc[idx]['open']
            sell_start_idx = idx + 1

        else:
            raise ValueError(f"不支持的买入模式:  {BUY_MODE}")

        # === 卖出逻辑 ===
        
        if SELL_MODE in ['open', 'close']:
            # 最多再拿 n_days‑1 根K线；第 HOLD_DAYS 根卖出
            sell_idx = sell_start_idx + HOLD_DAYS - 1
            if sell_idx >= len(df):        # 越界就跳过这笔交易
                continue
                
        if SELL_MODE == 'open':
            sell_price = df.iloc[sell_idx]['open']
            real_sell_idx = sell_idx
        elif SELL_MODE == 'close':
            sell_price = df.iloc[sell_idx]['close']
            real_sell_idx = sell_idx  
        elif SELL_MODE == 'strategy':
            max_days = MAX_HOLD_DAYS if MAX_HOLD_DAYS != -1 else len(df)

            if sell_start_idx >= len(df):
                continue  # 起始点越界，不能开始卖出窗口

            remaining_len = len(df) - sell_start_idx
            actual_window = min(max_days, remaining_len)

            if actual_window <= 0:
                continue

            all_sell_signals = sell_signal(df, [buy_date])

            # 防止返回空 DataFrame（可能是提前 return 的）
            if all_sell_signals is None or all_sell_signals.empty:
                continue

            sell_signals = all_sell_signals.iloc[sell_start_idx : sell_start_idx + actual_window]
            sell_window  = df.iloc[sell_start_idx : sell_start_idx + actual_window]

            if sell_signals.empty or sell_window.empty:
                continue

            sell_price = None
            for offset in range(0, len(sell_signals)):
                if sell_signals.iloc[offset]['sell_by_open']:
                    sell_price = sell_window.iloc[offset]['open']
                    real_sell_idx = sell_start_idx + offset
                    sell_start_idx += offset
                    break
                elif sell_signals.iloc[offset]['sell_by_close']:
                    sell_price = sell_window.iloc[offset]['close']
                    real_sell_idx = sell_start_idx + offset
                    sell_start_idx += offset
                    break

            if sell_price is None:
                if FALLBACK_SELL_MODE == 'open':
                    fallback_idx = sell_start_idx + HOLD_DAYS - 1
                    if fallback_idx < len(df):
                        sell_price = df.iloc[fallback_idx]['open']
                        real_sell_idx = fallback_idx
                    else:
                        continue  
                elif FALLBACK_SELL_MODE == 'close':
                    fallback_idx = sell_start_idx + HOLD_DAYS - 1
                    if fallback_idx < len(df):
                        sell_price = df.iloc[fallback_idx]['close']
                        real_sell_idx = fallback_idx
                    else:
                        continue
                else:
                    continue

        else:
            raise ValueError(f"不支持的 SELL_MODE: {SELL_MODE}")

        holding_until = df.index[real_sell_idx]

        # === 收益计算 ===
        ret = (sell_price - buy_price) / buy_price
        results.append(ret)
        if ret > 0:
            wins.append(ret)
        elif ret < 0:
            losses.append(ret)

#统计部分#################################################################################################
    returns = np.array(results)
    win_avg = np.mean(wins) if wins else 0
    loss_avg = abs(np.mean(losses)) if losses else 0
    win_rate = (returns > 0).mean() if len(results) else 0
    avg_return = returns.mean() if len(results) else 0
    if loss_avg == 0:
        pl_ratio = win_avg + 1
    else:
        pl_ratio = win_avg / loss_avg
    
    k = 5
    n = len(results)
    adjusted_pl_ratio = pl_ratio * n / (n + k)
    score = adjusted_pl_ratio * win_rate    

    if len(results) == 0:
        max_loss = None            # 或者写成 "N/A"
    elif len(results) == 1 and results[0] > 0:
        max_loss = 0
    else:
        max_loss = returns.min()

    summary = {
        "信号数": len(results),
        "平均收益": f"{avg_return:.2%}" if len(results) else None,
        "胜率": f"{win_rate:.2%}" if len(results) else None,
        "最大收益": f"{returns.max():.2%}" if len(results) else None,
        "最大亏损": f"{max_loss:.2%}" if max_loss is not None else None,
        "盈亏比": round(pl_ratio, 3) if len(results) else None,
        "平均盈亏比": round(score, 3) if len(results) else None
    }

    return summary


def run_backtest_on_file(file_path):
    df = pd.read_csv(file_path, parse_dates=['trade_date'])
    df.rename(columns={'trade_date': 'date'}, inplace=True)
    df = ensure_datetime_index(df)
    df = df[(df.index >= START_DATE) & (df.index <= END_DATE)]
    summary = backtest(
        df,
        HOLD_DAYS=HOLD_DAYS,
        BUY_MODE=BUY_MODE,
        SELL_MODE=SELL_MODE,
        MAX_HOLD_DAYS=MAX_HOLD_DAYS,
        FALLBACK_SELL_MODE=FALLBACK_SELL_MODE,
        buy_signal=buy_signal,
        sell_signal=sell_signal
    )
    return os.path.basename(file_path), summary

#主函数##################################################################################################
def main():
    start_time = time.time()
    
    all_results = []

    files = [os.path.join(DATA_DIR, f) for f in os.listdir(DATA_DIR) if f.endswith('.csv')]

    VERBOSE = 0

    with ProcessPoolExecutor() as executor:
        futures = {executor.submit(run_backtest_on_file, fpath): fpath for fpath in files}

        for future in tqdm(as_completed(futures), total=len(futures), desc="回测中", unit="支"):
            try:
                fname, summary = future.result()
                if VERBOSE:
                    print(f"【{fname}】结果：")
                    for k, v in summary.items():
                        print(f"  {k}: {v}")
                all_results.append(dict(filename=fname, **summary))
            except Exception as e:
                print(f"回测出错: {futures[future]} → {e}")

    end_time = time.time()

    # 保存 CSV/Excel 结果#################################################################################
    df_all = pd.DataFrame(all_results)
    
    valid_rows = df_all[pd.to_numeric(df_all["盈亏比"], errors='coerce').notna()]
    valid_rows.loc[:, "信号数"] = pd.to_numeric(valid_rows["信号数"], errors='coerce')
    valid_rows.loc[:, "盈亏比"] = pd.to_numeric(valid_rows["盈亏比"], errors='coerce')

    total_signals = valid_rows["信号数"].sum()
    if total_signals > 0:
        weighted_pl_ratio = (valid_rows["盈亏比"] * valid_rows["信号数"]).sum() / total_signals
    else:
        weighted_pl_ratio = 0

    df_all.to_excel(XLSX_PATH, index=False)

    # 添加条件格式
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    


    # 胜率 < 1 → G列标红
    ws.conditional_formatting.add("G2:G6000",
        CellIsRule(operator="lessThan", formula=["1"], stopIfTrue=True, fill=red_fill))
    # 平均盈亏比 < 1 → H列标红
    ws.conditional_formatting.add("H2:H6000",
        CellIsRule(operator="lessThan", formula=["1"], stopIfTrue=True, fill=red_fill))

    ws.insert_rows(1)
    ws.cell(row=1, column=1, value="加权平均盈亏比")
    ws.cell(row=1, column=2, value=round(weighted_pl_ratio, 4))

    wb.save(XLSX_PATH)
    
    print(f"\n总用时：{end_time - start_time:.2f} 秒")

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print("程序发生错误：", e)
        input("按下回车关闭窗口...")
